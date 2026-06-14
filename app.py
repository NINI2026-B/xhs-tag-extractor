import json
import io
import os
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import httpx

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "xhs-tag-extractor-secret-key-change-me")
PASSWORD = os.environ.get("PASSWORD", "168168")

# ══════ 登录保护 ══════
EXEMPT_ROUTES = ["/login", "/ping"]
LOGIN_HTML = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>登录 - 小红书标签提取</title>
<style>
  :root { --bg: #0f172a; --card: #1e293b; --text: #f1f5f9; --border: #334155; --accent: #FF2442; --accent-hover: #e61e3a; }
  * { margin:0; padding:0; box-sizing:border-box; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Noto Sans SC', sans-serif;
    background: var(--bg); color: var(--text);
    display: flex; justify-content: center; align-items: center; min-height: 100vh;
    background-image: radial-gradient(ellipse 80% 50% at 50% -20%, rgba(255,36,66,0.08), transparent);
  }
  .login-box {
    background: var(--card); border: 1px solid var(--border); border-radius: 16px;
    padding: 40px; width: 360px; text-align: center;
  }
  h1 { font-size: 1.3rem; margin-bottom: 24px; }
  h1 span { color: var(--accent); }
  input {
    width: 100%; padding: 12px 16px; border-radius: 8px; border: 1px solid var(--border);
    background: var(--bg); color: var(--text); font-size: 1.1rem; outline: none;
    margin-bottom: 16px; text-align: center; letter-spacing: 4px;
    transition: border-color 0.2s;
  }
  input:focus { border-color: var(--accent); box-shadow: 0 0 0 3px rgba(255,36,66,0.2); }
  button {
    width: 100%; padding: 12px; border: none; border-radius: 8px;
    background: var(--accent); color: white; font-size: 1rem; font-weight: 600;
    cursor: pointer; transition: all 0.2s;
  }
  button:hover { background: var(--accent-hover); }
  .error { color: #ef4444; font-size: 0.85rem; margin-top: 12px; display: none; }
</style>
</head>
<body>
<div class="login-box">
  <h1>🔐 <span>小红书标签提取</span></h1>
  <form method="POST" action="/login" onsubmit="return doLogin(event)">
    <input type="password" id="pwd" name="password" placeholder="请输入密码" autofocus>
    <button type="submit">登 录</button>
    <div class="error" id="error">密码错误，请重试</div>
  </form>
</div>
<script>
function doLogin(e) {
  e.preventDefault();
  var pwd = document.getElementById('pwd').value;
  fetch('/login', {method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({password:pwd})})
    .then(r => r.json()).then(d => {
      if (d.ok) location.href = '/';
      else document.getElementById('error').style.display = 'block';
    });
  return false;
}
</script>
</body>
</html>"""


@app.before_request
def check_auth():
    if request.path in EXEMPT_ROUTES:
        return None
    if session.get("logged_in"):
        return None
    if request.path == "/login":
        return None
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        data = request.get_json()
        pwd = data.get("password", "") if data else ""
        if pwd == PASSWORD:
            session["logged_in"] = True
            return jsonify({"ok": True})
        return jsonify({"ok": False}), 401
    return LOGIN_HTML


@app.route("/ping")
def ping():
    return "pong"

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"


def extract_tags_from_html(html):
    pos = html.find("window.__INITIAL_STATE__=")
    if pos == -1:
        return None, None, None

    json_start = pos + len("window.__INITIAL_STATE__=")
    brace_count = 0
    json_end = -1
    in_string = False
    escape = False
    for i in range(json_start, len(html)):
        c = html[i]
        if escape:
            escape = False
            continue
        if c == '\\':
            escape = True
            continue
        if c == '"':
            in_string = not in_string
            continue
        if in_string:
            continue
        if c == '{':
            brace_count += 1
        elif c == '}':
            brace_count -= 1
            if brace_count == 0:
                json_end = i + 1
                break

    if json_end == -1:
        return None, None, None

    json_str = html[json_start:json_end]
    json_str = json_str.replace("undefined", "null")

    try:
        data = json.loads(json_str)
    except json.JSONDecodeError:
        return None, None, None

    note_map = data.get("note", {}).get("noteDetailMap", {})
    if not note_map:
        return None, None, None

    first_note_id = list(note_map.keys())[0]
    note_data = note_map[first_note_id].get("note", {})

    tag_list = note_data.get("tagList", [])
    tags = [t["name"] for t in tag_list if t.get("name")]
    title = note_data.get("title", "")
    return title, tags, None


def fetch_one(client, url, idx):
    """抓取单个链接并提取标签，client 为 httpx.Client（线程安全）"""
    short_id = url.split("/")[-1] if "xhslink" in url else url[-40:]
    try:
        resp = client.get(url, headers={"User-Agent": UA})

        if resp.status_code != 200:
            return {"index": idx, "url": url, "shortId": short_id, "status": "failed", "title": "", "tags": [], "error": f"HTTP {resp.status_code}"}

        html = resp.text

        if "window.__INITIAL_STATE__=" not in html:
            return {"index": idx, "url": url, "shortId": short_id, "status": "failed", "title": "", "tags": [], "error": "页面不含 SSR 数据（可能需要登录或已失效）"}

        title, tags, _ = extract_tags_from_html(html)

        if tags is None:
            return {"index": idx, "url": url, "shortId": short_id, "status": "failed", "title": "", "tags": [], "error": "JSON 解析失败"}

        if not tags:
            return {"index": idx, "url": url, "shortId": short_id, "status": "failed", "title": title or "", "tags": [], "error": "未找到话题标签"}

        return {"index": idx, "url": url, "shortId": short_id, "status": "success", "title": title, "tags": tags}

    except httpx.TimeoutException:
        return {"index": idx, "url": url, "shortId": short_id, "status": "failed", "title": "", "tags": [], "error": "请求超时"}
    except Exception as e:
        return {"index": idx, "url": url, "shortId": short_id, "status": "failed", "title": "", "tags": [], "error": str(e)[:100]}


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/extract", methods=["POST"])
def extract():
    data = request.get_json()
    if not data or "urls" not in data:
        return jsonify({"error": "请提供 urls 参数"}), 400

    urls = data["urls"]
    if not isinstance(urls, list) or len(urls) == 0:
        return jsonify({"error": "urls 必须是非空数组"}), 400

    urls = urls[:100]

    results = []
    with httpx.Client(
        timeout=httpx.Timeout(20.0, connect=10.0),
        follow_redirects=True,
        limits=httpx.Limits(max_connections=10, max_keepalive_connections=5),
    ) as client:
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {executor.submit(fetch_one, client, url, i): i for i, url in enumerate(urls)}
            for future in as_completed(futures):
                results.append(future.result())

    results.sort(key=lambda x: x["index"])

    tag_counts = {}
    for r in results:
        if r["status"] == "success":
            for tag in r["tags"]:
                tag_counts[tag] = tag_counts.get(tag, 0) + 1

    stats = [{"tag": tag, "count": count} for tag, count in
             sorted(tag_counts.items(), key=lambda x: (-x[1], x[0]))]

    return jsonify({"results": results, "stats": stats})


@app.route("/api/export", methods=["POST"])
def export_excel():
    data = request.get_json()
    if not data or "results" not in data or "stats" not in data:
        return jsonify({"error": "请提供 results 和 stats 参数"}), 400

    results = data["results"]
    stats = data["stats"]

    wb = Workbook()

    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="FF2442", end_color="FF2442", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    cell_align = Alignment(vertical="center", wrap_text=True)

    ws1 = wb.active
    ws1.title = "提取结果"

    headers1 = ["序号", "链接ID", "状态", "标题", "话题标签", "错误信息"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, r in enumerate(results):
        row = i + 2
        values = [
            r["index"] + 1,
            r["shortId"],
            "成功" if r["status"] == "success" else "失败",
            r.get("title", ""),
            " ".join(["#" + t for t in r.get("tags", [])]),
            r.get("error", "") if r["status"] == "failed" else "",
        ]
        for col, v in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=v)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.font = Font(name="Microsoft YaHei", size=10)

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 8
    ws1.column_dimensions["D"].width = 40
    ws1.column_dimensions["E"].width = 60
    ws1.column_dimensions["F"].width = 25

    ws2 = wb.create_sheet("标签统计")

    headers2 = ["排名", "话题标签", "出现次数"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, s in enumerate(stats):
        row = i + 2
        values = [i + 1, "#" + s["tag"], s["count"]]
        for col, v in enumerate(values, 1):
            cell = ws2.cell(row=row, column=col, value=v)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.font = Font(name="Microsoft YaHei", size=10)

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"xiaohongshu_tags_{timestamp}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)